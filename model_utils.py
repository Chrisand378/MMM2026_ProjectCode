#%%
import math
from dataclasses import dataclass

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from config import (
    AGE_COL,
    CATEGORICAL_BACKGROUND_COLS,
    EDUCATION_COL,
    EXPERIENCE_COL,
    LOG_BACKGROUND_COLS,
    LOG_BACKGROUND_PREFIX,
    LOG_WAGE_COL,
    MODEL_EDUCATION_COL,
    NUMERIC_BACKGROUND_COLS,
    WAGE_COL,
)


#%%
def normal_cdf(x):
    """Standard normal CDF"""
    x = np.asarray(x, dtype=float)
    return 0.5 * (1.0 + np.vectorize(math.erf)(x / math.sqrt(2.0)))


def add_constant(x, name="const"):
    x = pd.DataFrame(x).copy()
    if name not in x.columns:
        x.insert(0, name, 1.0)
    return x


def standardize(series):
    values = pd.Series(series, dtype=float)
    std = values.std(ddof=0)
    if std == 0 or np.isnan(std):
        return values * 0.0
    return (values - values.mean()) / std


def safe_log(values, eps=1e-8):
    return np.log(np.clip(np.asarray(values, dtype=float), eps, None))


#%%
def preprocess_individual_data(data, require_wage=False):
    """Apply project-wide dataset rules."""
    data = data.copy()

    data[MODEL_EDUCATION_COL] = pd.to_numeric(data[EDUCATION_COL]).astype(int)

    if require_wage:
        wage = pd.to_numeric(data[WAGE_COL], errors="coerce")
        data[LOG_WAGE_COL] = np.nan
        data.loc[wage > 0, LOG_WAGE_COL] = np.log(wage.loc[wage > 0])

    for col in [AGE_COL, EXPERIENCE_COL] + NUMERIC_BACKGROUND_COLS + LOG_BACKGROUND_COLS:
        data[col] = pd.to_numeric(data[col], errors="coerce")

    for col in LOG_BACKGROUND_COLS:
        data[f"{LOG_BACKGROUND_PREFIX}{col}"] = np.log1p(data[col].clip(lower=0))

    return data


def build_background_regressors(data):
    """Build background controls."""
    numeric_cols = NUMERIC_BACKGROUND_COLS + [
        f"{LOG_BACKGROUND_PREFIX}{col}" for col in LOG_BACKGROUND_COLS
    ]

    parts = [data[numeric_cols].astype(float)]

    for col in CATEGORICAL_BACKGROUND_COLS:
        dummies = pd.get_dummies(
            data[col].astype("category"),
            prefix=col,
            drop_first=True,
            dtype=float,
        )
        parts.append(dummies)

    return pd.concat(parts, axis=1)


#%%
def load_named_excel_table(path, table_name):
    """Read an Excel named table into a pandas DataFrame."""
    workbook = load_workbook(path, data_only=True)
    for worksheet in workbook.worksheets:
        if table_name not in worksheet.tables:
            continue

        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        rows = list(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )
        return pd.DataFrame(rows[1:], columns=rows[0])

    raise ValueError(f"Could not find table '{table_name}' in {path}")


def load_individual_csv(path):
    """Read the individual-level CSV."""
    return pd.read_csv(path)


#%%
def _two_sided_p_values(t_values, df_resid):
    """
    Two-sided p-values for coefficient tests.

    Uses scipy's Student-t distribution when available. If scipy is not
    installed, falls back to the large-sample normal approximation.
    """
    t_values = np.asarray(t_values, dtype=float)
    with np.errstate(invalid="ignore"):
        abs_t = np.abs(t_values)

    try:
        from scipy import stats  # optional dependency

        return 2.0 * stats.t.sf(abs_t, df=max(int(df_resid), 1))
    except Exception:
        return 2.0 * (1.0 - normal_cdf(abs_t))


@dataclass
class OLSResult:
    coefficients: pd.Series
    standard_errors: pd.Series
    t_values: pd.Series
    p_values: pd.Series
    residual_std: float
    r_squared: float
    log_likelihood: float
    aic: float
    bic: float
    nobs: int
    n_params: int
    df_resid: int
    fitted: pd.Series
    residuals: pd.Series


def fit_ols(y, x):
    """OLS with classical standard errors, t-values, p-values, and Gaussian log-likelihood."""
    y = pd.Series(y, dtype=float).reset_index(drop=True)
    x = add_constant(x).astype(float).reset_index(drop=True)

    keep = y.notna() & x.notna().all(axis=1)
    y = y.loc[keep].reset_index(drop=True)
    x = x.loc[keep].reset_index(drop=True)

    x_mat = x.to_numpy()
    y_vec = y.to_numpy()
    nobs, n_params = x_mat.shape
    df_resid = max(nobs - n_params, 1)

    beta = np.linalg.pinv(x_mat.T @ x_mat) @ x_mat.T @ y_vec
    fitted = x_mat @ beta
    resid = y_vec - fitted

    sse = float(resid.T @ resid)
    tss = float(((y_vec - y_vec.mean()) ** 2).sum())
    sigma2 = sse / df_resid
    cov = sigma2 * np.linalg.pinv(x_mat.T @ x_mat)
    se = np.sqrt(np.diag(cov))

    with np.errstate(divide="ignore", invalid="ignore"):
        t_values = beta / se
    p_values = _two_sided_p_values(t_values, df_resid)

    ll_sigma2 = max(sse / max(nobs, 1), 1e-12)
    log_likelihood = float(-0.5 * nobs * (math.log(2 * math.pi) + math.log(ll_sigma2) + 1))

    return OLSResult(
        coefficients=pd.Series(beta, index=x.columns),
        standard_errors=pd.Series(se, index=x.columns),
        t_values=pd.Series(t_values, index=x.columns),
        p_values=pd.Series(p_values, index=x.columns),
        residual_std=float(math.sqrt(sigma2)),
        r_squared=float(1 - sse / tss) if tss > 0 else np.nan,
        log_likelihood=log_likelihood,
        aic=float(2 * n_params - 2 * log_likelihood),
        bic=float(math.log(max(nobs, 1)) * n_params - 2 * log_likelihood),
        nobs=int(nobs),
        n_params=int(n_params),
        df_resid=int(df_resid),
        fitted=pd.Series(fitted),
        residuals=pd.Series(resid),
    )



#%%
def print_model_summary(name, result):
    """Compact console summary for interactive runs."""
    print(f"\n{name}")
    print("=" * len(name))
    if hasattr(result, "coefficients"):
        print("Coefficients")
        print(result.coefficients)
    if hasattr(result, "standard_errors"):
        print("\nStandard errors")
        print(result.standard_errors)
    if hasattr(result, "p_values"):
        print("\nP-values")
        print(result.p_values)
    for attr in ["nobs", "r_squared", "residual_std", "log_likelihood", "aic", "bic"]:
        if hasattr(result, attr):
            print(f"{attr}: {getattr(result, attr)}")
